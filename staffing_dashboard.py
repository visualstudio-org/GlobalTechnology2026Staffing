"""
Global Technology 2026 Staffing Rampup Plan Dashboard
Interactive dashboard for tracking technology team hiring and resource allocation
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import numpy as np

# Page configuration
st.set_page_config(
    page_title="First Advantage | Tech Staffing 2026",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded"
)

# First Advantage Brand Colors (from Tech in Review)
FA_GREEN = "#00a84f"  # Primary Green
FA_GREEN_DARK = "#006838"  # Dark Green
FA_GREEN_LIGHT = "#4dc47d"  # Light Green
FA_NAVY = "#1a1a2e"  # Navy
FA_DARK = "#2d2d2d"
FA_GRAY = "#6b7280"
FA_LIGHT_GRAY = "#f8f9fa"
FA_WARNING = "#f5a623"  # Orange/Gold accent

# Custom CSS with First Advantage Branding
st.markdown(f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700&display=swap');
    
    /* Global Styles */
    .main {{
        background-color: #FFFFFF;
        font-family: 'Inter', sans-serif;
    }}
    
    /* Metrics Styling */
    [data-testid="stMetricValue"] {{
        font-size: 2rem;
        font-weight: 700;
        color: {FA_GREEN};
    }}
    
    [data-testid="stMetricLabel"] {{
        font-size: 0.9rem;
        font-weight: 600;
        color: {FA_GRAY};
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }}
    
    [data-testid="stMetricDelta"] {{
        font-weight: 600;
    }}
    
    /* Sidebar Styling */
    [data-testid="stSidebar"] {{
        background: linear-gradient(180deg, {FA_LIGHT_GRAY} 0%, #FFFFFF 100%);
        border-right: 3px solid {FA_GREEN};
    }}
    
    [data-testid="stSidebar"] .stRadio label {{
        font-weight: 600;
        color: {FA_NAVY};
    }}
    
    /* Radio Button Styling - Green */
    input[type="radio"] {{
        accent-color: {FA_GREEN} !important;
    }}
    
    input[type="radio"]:checked {{
        accent-color: {FA_GREEN} !important;
        background-color: {FA_GREEN} !important;
    }}
    
    .stRadio > label > div[role="radiogroup"] > label > div:first-child {{
        background-color: {FA_GREEN} !important;
    }}
    
    /* Multiselect Filter Styling - Green */
    [data-testid="stMultiSelect"] {{
        color: {FA_GREEN};
    }}
    
    [data-testid="stMultiSelect"] > div > div {{
        border-color: {FA_GREEN};
    }}
    
    [data-testid="stMultiSelect"] span[data-baseweb="tag"] {{
        background-color: {FA_GREEN} !important;
        color: white !important;
    }}
    
    [data-testid="stMultiSelect"] span[data-baseweb="tag"] button {{
        color: white !important;
    }}
    
    /* Tabs Styling */
    .stTabs [data-baseweb="tab-list"] {{
        gap: 8px;
        background-color: {FA_LIGHT_GRAY};
        padding: 10px;
        border-radius: 10px;
    }}
    
    .stTabs [data-baseweb="tab"] {{
        height: 50px;
        background-color: white;
        border-radius: 8px;
        color: {FA_NAVY};
        font-weight: 600;
        border: 2px solid transparent;
        transition: all 0.3s ease;
    }}
    
    .stTabs [aria-selected="true"] {{
        background: linear-gradient(135deg, {FA_GREEN_DARK} 0%, {FA_GREEN} 100%);
        color: white;
        border: 2px solid {FA_WARNING};
    }}
    
    /* Button Styling */
    .stButton > button {{
        background: linear-gradient(135deg, {FA_GREEN_DARK} 0%, {FA_GREEN} 100%);
        color: white;
        font-weight: 600;
        border: none;
        border-radius: 8px;
        padding: 10px 24px;
        transition: all 0.3s ease;
        box-shadow: 0 2px 8px rgba(0, 168, 79, 0.3);
    }}
    
    .stButton > button:hover {{
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0, 168, 79, 0.4);
        background: linear-gradient(135deg, {FA_GREEN} 0%, {FA_GREEN_LIGHT} 100%);
    }}
    
    /* Download Button */
    .stDownloadButton > button {{
        background-color: {FA_WARNING};
        color: black !important;
        font-weight: 600;
        border-radius: 8px;
        border: none;
        padding: 10px 24px;
        transition: all 0.3s ease;
    }}
    
    .stDownloadButton > button:hover {{
        background-color: #E59416;
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(245, 166, 35, 0.4);
        color: black !important;
    }}
    
    /* Progress Bar */
    .stProgress > div > div > div > div {{
        background: linear-gradient(90deg, {FA_GREEN_DARK} 0%, {FA_GREEN} 50%, {FA_GREEN_LIGHT} 100%);
    }}
    
    /* Dataframe Styling */
    [data-testid="stDataFrame"] {{
        border: 2px solid {FA_LIGHT_GRAY};
        border-radius: 10px;
        overflow: hidden;
    }}
    
    /* Section Headers */
    h2, h3 {{
        color: {FA_GREEN};
        font-weight: 700;
        margin-top: 2rem;
    }}
    
    /* Info Box */
    .stAlert {{
        border-radius: 10px;
        border-left: 4px solid {FA_GREEN};
    }}
    
    /* Footer */
    .footer {{
        text-align: center;
        padding: 20px;
        color: {FA_GRAY};
        font-size: 0.9rem;
        border-top: 2px solid {FA_LIGHT_GRAY};
        margin-top: 40px;
    }}
    </style>
""", unsafe_allow_html=True)

# File path
FILE_PATH = r'C:\Users\Eric.Jaffe\OneDrive - First Advantage Corporation\2026 Budget\Global Technology 2026 Staffing Rampup Plan 011226.xlsx'

@st.cache_data
def load_data():
    """Load data from Excel file"""
    try:
        # Load summary data
        summary_df = pd.read_excel(FILE_PATH, sheet_name='Technology Staffing Summary', header=1)
        if '#' in summary_df.columns:
            summary_df = summary_df[summary_df['#'].notna()]  # Remove empty rows
        
        # Convert numeric columns to proper types
        numeric_cols = ['# of New Roles', 'Est. Investment', 'Open Roles', 'Closed Roles']
        for col in numeric_cols:
            if col in summary_df.columns:
                summary_df[col] = pd.to_numeric(summary_df[col], errors='coerce').fillna(0)
        
        # Load detailed data - header is in row 2 (0-indexed row 1)
        detailed_df = pd.read_excel(FILE_PATH, sheet_name='Detailed 2026 Staffing Plans', header=1, skiprows=[0])
        # Remove empty rows
        if 'Technology Area' in detailed_df.columns:
            detailed_df = detailed_df[detailed_df['Technology Area'].notna()].copy()
            # Convert date columns to strings to avoid conversion issues
            date_cols = ['Target \\nStart Date', 'Target \\nEnd Date', 'Actual Start', 'Actual End Date']
            for col in date_cols:
                if col in detailed_df.columns:
                    detailed_df[col] = detailed_df[col].astype(str)
        
        return summary_df, detailed_df
    except Exception as e:
        st.error(f"Error loading file: {e}")
        st.info("Please make sure the Excel file is closed and try refreshing the page.")
        import traceback
        st.error(traceback.format_exc())
        return None, None

def main():
    # Auto-refresh data every 15 minutes
    if 'last_refresh' not in st.session_state:
        st.session_state.last_refresh = datetime.now()
    
    # Check if 15 minutes have passed
    time_since_refresh = (datetime.now() - st.session_state.last_refresh).total_seconds()
    if time_since_refresh > 900:  # 900 seconds = 15 minutes
        st.cache_data.clear()
        st.session_state.last_refresh = datetime.now()
        st.rerun()
    
    # Header with First Advantage Branding
    col1, col2 = st.columns([1, 5])
    with col1:
        try:
            st.image(r"C:\Users\Eric.Jaffe\OneDrive - First Advantage Corporation\Desktop\partner-FirstAdvantage-logo-1.png", width=150)
        except:
            pass  # Hide fallback text
    
    with col2:
        st.markdown(f"""
            <div style='padding-top: 20px;'>
                <h1 style='color: {FA_GREEN}; margin-bottom: 0; font-weight: 700;'>Global Technology 2026 Staffing Dashboard</h1>
            </div>
        """, unsafe_allow_html=True)
    
    st.markdown("<hr style='margin-top: 10px; margin-bottom: 30px; border-color: #00a84f;'>", unsafe_allow_html=True)
    
    # Load data
    summary_df, detailed_df = load_data()
    
    if summary_df is None:
        return
    
    # Sidebar
    st.sidebar.title("📊 Dashboard Controls")
    st.sidebar.info(f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    
    # Add refresh button
    if st.sidebar.button("🔄 Refresh Data"):
        st.cache_data.clear()
        st.rerun()
    
    # Filter by Technology Area or Investment Area
    filter_type = st.sidebar.radio("View By:", ["Technology Area", "Investment Area"])
    
    # Split data into Technology Areas and Investment Areas
    # Only use rows where 'Technology Area' column has actual technology areas (not investment areas)
    # The first section contains Technology Areas, then there's a break, then Investment Areas
    tech_areas_df = summary_df[
        (summary_df['Technology Area'].notna()) & 
        (summary_df['#'].notna()) &
        (~summary_df['Technology Area'].str.contains('Investment', case=False, na=False))
    ].head(6).copy()  # Limit to first 6 technology areas to avoid duplicates
    
    # Get investment areas (rows where 'Investment Area' column is populated)
    investment_df = summary_df[
        (summary_df['Investment Area'].notna() if 'Investment Area' in summary_df.columns else False)
    ].copy() if 'Investment Area' in summary_df.columns else pd.DataFrame()
    
    # Key Metrics Row
    st.subheader("📈 Overall Metrics")
    col1, col2, col3, col4, col5 = st.columns(5)
    
    total_roles = tech_areas_df['# of New Roles'].sum()
    total_investment = tech_areas_df['Est. Investment'].sum()
    total_open = tech_areas_df['Open Roles'].sum()
    total_closed = tech_areas_df['Closed Roles'].sum()
    close_rate = (total_closed / total_roles * 100) if total_roles > 0 else 0
    
    with col1:
        st.metric("Total New Roles", f"{int(total_roles)}")
    
    with col2:
        st.metric("Total Investment", f"${total_investment/1000000:.2f}M")
    
    with col3:
        st.metric("Open Roles", f"{int(total_open)}")
    
    with col4:
        st.metric("Closed Roles", f"{int(total_closed)}",
                 delta=f"{close_rate:.1f}%")
    
    with col5:
        avg_cost = total_investment / total_roles if total_roles > 0 else 0
        st.metric("Avg Cost/Role", f"${avg_cost/1000:.0f}K")
    
    # Tabs
    tab1, tab2, tab3, tab4 = st.tabs([
        "📊 Overview",
        "🎯 Technology Areas", 
        "💰 Investment Analysis",
        "📋 Detailed Data"
    ])
    
    with tab1:
        st.subheader("Hiring Progress Overview")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Overall progress
            fig1 = go.Figure(go.Indicator(
                mode = "gauge+number+delta",
                value = total_closed,
                delta = {'reference': total_roles},
                title = {'text': f"Roles Filled ({close_rate:.1f}%)", 'font': {'size': 20, 'color': FA_GREEN}},
                gauge = {
                    'axis': {'range': [None, total_roles]},
                    'bar': {'color': FA_GREEN_LIGHT},
                    'steps': [
                        {'range': [0, total_roles*0.5], 'color': FA_LIGHT_GRAY},
                        {'range': [total_roles*0.5, total_roles*0.8], 'color': '#D3D3D3'}
                    ],
                    'threshold': {
                        'line': {'color': FA_WARNING, 'width': 4},
                        'thickness': 0.75,
                        'value': total_roles
                    }
                }
            ))
            fig1.update_layout(height=300, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
            st.plotly_chart(fig1, use_container_width=True)
        
        with col2:
            # Status breakdown
            status_data = pd.DataFrame({
                'Status': ['Closed (Filled)', 'Open (Recruiting)'],
                'Count': [total_closed, total_open]
            })
            fig2 = px.pie(status_data, values='Count', names='Status',
                         title='Recruitment Status',
                         color_discrete_sequence=[FA_GREEN_LIGHT, FA_WARNING])
            fig2.update_layout(height=300, paper_bgcolor='rgba(0,0,0,0)', 
                             title_font_color=FA_GREEN, title_font_size=16)
            st.plotly_chart(fig2, use_container_width=True)
        
        # Technology Areas Overview
        st.subheader("Technology Areas Breakdown")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Roles by Technology Area
            tech_sorted = tech_areas_df.sort_values('# of New Roles', ascending=True)
            fig3 = px.bar(tech_sorted, 
                         y='Technology Area', 
                         x='# of New Roles',
                         title='New Roles by Technology Area',
                         orientation='h',
                         color='# of New Roles',
                         color_continuous_scale=[[0, FA_GREEN_LIGHT], [1, FA_GREEN]])
            fig3.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                             title_font_color=FA_GREEN, title_font_size=16)
            st.plotly_chart(fig3, use_container_width=True)
        
        with col2:
            # Investment by Technology Area
            fa_colors = [FA_GREEN, FA_GREEN_LIGHT, FA_WARNING, '#4A90E2', '#F39C12', '#8E44AD']
            fig4 = px.pie(tech_areas_df, 
                         values='Est. Investment', 
                         names='Technology Area',
                         title='Investment Distribution by Technology Area',
                         color_discrete_sequence=fa_colors)
            fig4.update_layout(paper_bgcolor='rgba(0,0,0,0)',
                             title_font_color=FA_GREEN, title_font_size=16)
            st.plotly_chart(fig4, use_container_width=True)
    
    with tab2:
        st.subheader("Technology Area Details")
        
        # Select technology area
        tech_area = st.selectbox("Select Technology Area", 
                                tech_areas_df['Technology Area'].unique())
        
        if tech_area:
            selected_tech = tech_areas_df[tech_areas_df['Technology Area'] == tech_area].iloc[0]
            
            # Show metrics for selected area
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("Leader", selected_tech['Leaders'])
            with col2:
                st.metric("Total Roles", int(selected_tech['# of New Roles']))
            with col3:
                st.metric("Open", int(selected_tech['Open Roles']))
            with col4:
                st.metric("Closed", int(selected_tech['Closed Roles']))
            
            # Progress bar
            total_roles = selected_tech['# of New Roles']
            if total_roles > 0:
                progress = selected_tech['Closed Roles'] / total_roles
                st.progress(progress, text=f"Progress: {progress*100:.1f}%")
            else:
                st.progress(0.0, text="Progress: 0.0%")
            
            # Investment details
            st.metric("Estimated Investment", 
                     f"${selected_tech['Est. Investment']/1000000:.2f}M")
        
        # Comparison chart
        st.subheader("Technology Area Comparison")
        
        comparison_df = tech_areas_df.copy()
        comparison_df['Close Rate %'] = (comparison_df['Closed Roles'] / 
                                          comparison_df['# of New Roles'] * 100)
        
        fig5 = go.Figure()
        fig5.add_trace(go.Bar(
            x=comparison_df['Technology Area'],
            y=comparison_df['Open Roles'],
            name='Open Roles',
            marker_color=FA_WARNING
        ))
        fig5.add_trace(go.Bar(
            x=comparison_df['Technology Area'],
            y=comparison_df['Closed Roles'],
            name='Closed Roles',
            marker_color=FA_GREEN_LIGHT
        ))
        fig5.update_layout(
            barmode='stack',
            title='Recruitment Progress by Technology Area',
            xaxis_title='Technology Area',
            yaxis_title='Number of Roles',
            hovermode='x unified',
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)',
            title_font_color=FA_GREEN,
            title_font_size=16
        )
        st.plotly_chart(fig5, use_container_width=True)
        
        # Close rate comparison
        fig6 = px.bar(comparison_df,
                     x='Technology Area',
                     y='Close Rate %',
                     title='Close Rate by Technology Area',
                     color='Close Rate %',
                     color_continuous_scale=[[0, FA_WARNING], [0.5, FA_GREEN_LIGHT], [1, FA_GREEN]],
                     text='Close Rate %')
        fig6.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
        fig6.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                         title_font_color=FA_GREEN, title_font_size=16)
        st.plotly_chart(fig6, use_container_width=True)
    
    with tab3:
        st.subheader("Investment Analysis")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Investment vs Roles  
            fig7 = px.scatter(tech_areas_df,
                            x='# of New Roles',
                            y='Est. Investment',
                            size='# of New Roles',
                            color='Technology Area',
                            hover_data=['Leaders'],
                            title='Investment vs Number of Roles',
                            labels={'Est. Investment': 'Investment ($)',
                                   '# of New Roles': 'Number of Roles'},
                            color_discrete_sequence=[FA_GREEN, FA_GREEN_LIGHT, FA_WARNING, '#4A90E2', '#F39C12', '#8E44AD'])
            fig7.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                             title_font_color=FA_GREEN, title_font_size=16)
            st.plotly_chart(fig7, width='stretch')
        
        with col2:
            # Average cost per role
            cost_df = tech_areas_df.copy()
            cost_df['Avg Cost per Role'] = cost_df['Est. Investment'] / cost_df['# of New Roles']
            
            fig8 = px.bar(cost_df.sort_values('Avg Cost per Role', ascending=False),
                         x='Technology Area',
                         y='Avg Cost per Role',
                         title='Average Cost per Role by Technology Area',
                         color='Avg Cost per Role',
                         color_continuous_scale=[[0, FA_GREEN_LIGHT], [1, FA_GREEN]])
            fig8.update_layout(xaxis_tickangle=-45, paper_bgcolor='rgba(0,0,0,0)', 
                             plot_bgcolor='rgba(0,0,0,0)', title_font_color=FA_GREEN, title_font_size=16)
            st.plotly_chart(fig8, use_container_width=True)
        
        # Investment breakdown table
        st.subheader("Investment Breakdown")
        
        invest_summary = tech_areas_df[['Technology Area', 'Leaders', '# of New Roles', 
                                        'Est. Investment', 'Open Roles', 'Closed Roles']].copy()
        invest_summary['Avg Cost/Role'] = invest_summary['Est. Investment'] / invest_summary['# of New Roles']
        invest_summary['Close Rate %'] = (invest_summary['Closed Roles'] / 
                                          invest_summary['# of New Roles'] * 100)
        
        # Format currency columns
        invest_summary['Est. Investment'] = invest_summary['Est. Investment'].apply(lambda x: f'${x:,.0f}')
        invest_summary['Avg Cost/Role'] = invest_summary['Avg Cost/Role'].apply(lambda x: f'${x:,.0f}')
        invest_summary['Close Rate %'] = invest_summary['Close Rate %'].apply(lambda x: f'{x:.1f}%')
        
        st.dataframe(invest_summary, use_container_width=True, height=400)
    
    with tab4:
        st.subheader("Detailed Staffing Data")
        
        # Display raw summary data
        st.write("### Technology Areas Summary")
        display_df = tech_areas_df[['#', 'Technology Area', 'Leaders', '# of New Roles', 
                                    'Est. Investment', 'Open Roles', 'Closed Roles']].copy()
        # Format Est. Investment as currency
        display_df['Est. Investment'] = display_df['Est. Investment'].apply(lambda x: f'${x:,.0f}')
        st.dataframe(display_df, use_container_width=True)
        
        # Display detailed roles data
        st.write("### Detailed Roles Breakdown")
        
        if detailed_df is not None and len(detailed_df) > 0:
            # Add filters for detailed data
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if 'Status' in detailed_df.columns:
                    status_options = ['All'] + sorted([s for s in detailed_df['Status'].dropna().unique() if s])
                    status_filter = st.multiselect(
                        "Filter by Status",
                        options=status_options,
                        default=['All']
                    )
                else:
                    status_filter = ['All']
            
            with col2:
                if 'Technology Area' in detailed_df.columns:
                    tech_options = ['All'] + sorted([t for t in detailed_df['Technology Area'].dropna().unique() if t])
                    tech_filter = st.multiselect(
                        "Filter by Technology Area",
                        options=tech_options,
                        default=['All']
                    )
                else:
                    tech_filter = ['All']
            
            with col3:
                if 'TEAM NAME' in detailed_df.columns:
                    team_options = ['All'] + sorted([t for t in detailed_df['TEAM NAME'].dropna().unique() if t])
                    team_filter = st.multiselect(
                        "Filter by Team",
                        options=team_options,
                        default=['All']
                    )
                else:
                    team_filter = ['All']
            
            # Apply filters
            filtered_detailed = detailed_df.copy()
            
            if 'Status' in detailed_df.columns and 'All' not in status_filter and len(status_filter) > 0:
                filtered_detailed = filtered_detailed[filtered_detailed['Status'].isin(status_filter)]
            
            if 'Technology Area' in detailed_df.columns and 'All' not in tech_filter and len(tech_filter) > 0:
                filtered_detailed = filtered_detailed[filtered_detailed['Technology Area'].isin(tech_filter)]
            
            if 'TEAM NAME' in detailed_df.columns and 'All' not in team_filter and len(team_filter) > 0:
                filtered_detailed = filtered_detailed[filtered_detailed['TEAM NAME'].isin(team_filter)]
            
            # Show metrics for filtered data
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Filtered Roles", len(filtered_detailed))
            with col2:
                # Count all non-Closed roles as Open
                closed_count = len(filtered_detailed[filtered_detailed['Status'] == 'Closed']) if 'Status' in filtered_detailed.columns else 0
                open_count = len(filtered_detailed) - closed_count
                st.metric("Open", open_count)
            with col3:
                st.metric("Closed", closed_count)
            
            # Select key columns to display in specified order
            display_cols = ['Technology Area', 'TEAM NAME', 'Worker Type', 'Req ID', 
                          'Recruitment Status', 'Location', 'Senior Leader', 'Hiring Manager',
                          'Target \nStart Date', 'Target \nEnd Date', 'Actual Start', 'Status', 'Comment']
            available_cols = [col for col in display_cols if col in filtered_detailed.columns]
            
            # Format display data
            display_data = filtered_detailed[available_cols].copy()
            
            # Rename columns to remove newlines for better display
            column_rename = {
                'Target \nStart Date': 'Target Start Date',
                'Target \nEnd Date': 'Target End Date'
            }
            display_data = display_data.rename(columns=column_rename)
            
            # Format date columns as MM/DD/YYYY
            date_columns = ['Target Start Date', 'Target End Date', 'Actual Start']
            for date_col in date_columns:
                if date_col in display_data.columns:
                    display_data[date_col] = pd.to_datetime(display_data[date_col], errors='coerce').dt.strftime('%m/%d/%Y')
                    display_data[date_col] = display_data[date_col].replace('NaT', '')
            
            # Display the detailed data (read-only for now to prevent data corruption)
            st.write("**Note:** Data is read-only. Edit the Excel file directly in OneDrive for updates.")
            
            st.dataframe(
                display_data, 
                use_container_width=True, 
                height=400
            )
        else:
            st.info("No detailed roles data available.")
        
        # Export options
        st.subheader("📥 Export Data")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Create formatted Excel for Summary
            from io import BytesIO
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Remove hidden columns
            summary_export = display_df.copy()
            cols_to_hide = ['Est. Blended Hourly Rate', 'Est. Forecast']
            summary_export = summary_export.drop(columns=[col for col in cols_to_hide if col in summary_export.columns], errors='ignore')
            
            # Format column headers - convert date timestamps to MM-DD-YY format for columns Z onwards
            new_columns = []
            for col in summary_export.columns:
                col_str = str(col)
                # Try to parse as date and format if it looks like a date/timestamp
                try:
                    # Remove any ISO timestamp portion (e.g., "2025-12-09T15:45:19z")
                    if 'T' in col_str or len(col_str) > 10:
                        date_part = col_str.split('T')[0]  # Get date portion before 'T'
                        parsed_date = pd.to_datetime(date_part)
                        new_columns.append(parsed_date.strftime('%m-%d-%y'))
                    else:
                        # Try parsing as regular date
                        parsed_date = pd.to_datetime(col_str)
                        new_columns.append(parsed_date.strftime('%m-%d-%y'))
                except:
                    # Not a date, keep original
                    new_columns.append(col)
            summary_export.columns = new_columns
            
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                summary_export.to_excel(writer, sheet_name='Summary', index=False)
                workbook = writer.book
                worksheet = writer.sheets['Summary']
                
                # Format styles
                header_fill = PatternFill(start_color='00A84F', end_color='00A84F', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=12)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Format header row and enable auto-filter
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                
                # Enable auto-filter on all columns
                worksheet.auto_filter.ref = worksheet.dimensions
                
                # Format all data cells with borders and date formatting
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = thin_border
                        # Format dates as mm/dd/yyyy
                        if cell.value and isinstance(cell.value, (datetime, pd.Timestamp)):
                            cell.number_format = 'mm/dd/yyyy'
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            st.download_button(
                label="📊 Download Summary as Excel",
                data=buffer.getvalue(),
                file_name=f"tech_staffing_summary_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        with col2:
            if detailed_df is not None and len(detailed_df) > 0:
                # Create formatted Excel for Detailed Roles
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                # Remove hidden columns
                detailed_export = detailed_df.copy()
                cols_to_hide = ['Est. Blended Hourly Rate', 'Est. Forecast']
                detailed_export = detailed_export.drop(columns=[col for col in cols_to_hide if col in detailed_export.columns], errors='ignore')
                
                # Format column headers - convert date timestamps to MM-DD-YY format for columns Z onwards
                new_columns = []
                for col in detailed_export.columns:
                    col_str = str(col)
                    # Try to parse as date and format if it looks like a date/timestamp
                    try:
                        # Remove any ISO timestamp portion (e.g., "2025-12-09T15:45:19z")
                        if 'T' in col_str or len(col_str) > 10:
                            date_part = col_str.split('T')[0]  # Get date portion before 'T'
                            parsed_date = pd.to_datetime(date_part)
                            new_columns.append(parsed_date.strftime('%m-%d-%y'))
                        else:
                            # Try parsing as regular date
                            parsed_date = pd.to_datetime(col_str)
                            new_columns.append(parsed_date.strftime('%m-%d-%y'))
                    except:
                        # Not a date, keep original
                        new_columns.append(col)
                detailed_export.columns = new_columns
                
                buffer2 = BytesIO()
                with pd.ExcelWriter(buffer2, engine='openpyxl') as writer:
                    detailed_export.to_excel(writer, sheet_name='Detailed Roles', index=False)
                    workbook = writer.book
                    worksheet = writer.sheets['Detailed Roles']
                    
                    # Format styles
                    header_fill = PatternFill(start_color='00A84F', end_color='00A84F', fill_type='solid')
                    header_font = Font(bold=True, color='FFFFFF', size=12)
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Format header row
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                    
                    # Enable auto-filter on all columns
                    worksheet.auto_filter.ref = worksheet.dimensions
                    
                    # Format all data cells with borders and date formatting
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                        for cell in row:
                            cell.border = thin_border
                            # Format dates as mm/dd/yyyy
                            if cell.value and isinstance(cell.value, (datetime, pd.Timestamp)):
                                cell.number_format = 'mm/dd/yyyy'
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
                st.download_button(
                    label="📋 Download Detailed Roles as Excel",
                    data=buffer2.getvalue(),
                    file_name=f"tech_staffing_detailed_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                # Summary report
                if st.button("📄 Generate Summary Report"):
                    st.write("### Summary Statistics")
                    st.write(f"- Total Technology Areas: {len(tech_areas_df)}")
                    st.write(f"- Total New Roles: {int(total_roles)}")
                    st.write(f"- Total Investment: ${total_investment:,.0f}")
                    st.write(f"- Average Investment per Role: ${avg_cost:,.0f}")
                    st.write(f"- Overall Close Rate: {close_rate:.1f}%")
                    st.write(f"- Roles Still Open: {int(total_open)}")
    
    # Footer
    st.markdown("---")
    st.markdown(f"""
        <div class="footer">
            <p><strong>FIRST ADVANTAGE</strong> | Global Technology 2026 Staffing Dashboard</p>
            <p>📅 Last Updated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</p>
            <p style='font-size: 0.8rem; color: {FA_GRAY};'>Data Source: Global Technology 2026 Staffing Rampup Plan</p>
        </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
